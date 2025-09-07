from django.contrib.admin import SimpleListFilter
# Filtro mês/ano para o campo 'data' de Entrada
class MesAnoDataEntradaFilter(SimpleListFilter):
    title = 'Mês/Ano'
    parameter_name = 'mes_ano'

    def lookups(self, request, model_admin):
        datas = model_admin.model.objects.order_by('-data').values_list('data', flat=True)
        opcoes = set()
        for data in datas:
            opcoes.add((data.strftime('%Y-%m'), data.strftime('%m/%Y')))
        return sorted(opcoes, reverse=True)

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            ano, mes = value.split('-')
            return queryset.filter(data__year=ano, data__month=mes)
        return queryset
# Action para exportar entradas selecionadas para XLSX
def exportar_entradas_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Descrição', 'Valor', 'Data', 'Categoria', 'Recebido por'])
    for obj in queryset:
        ws.append([
            obj.descricao,
            obj.valor,
            obj.data.strftime('%d/%m/%Y'),
            obj.categoria,
            obj.recebido_por
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=entradas.xlsx'
    wb.save(response)
    return response

exportar_entradas_xlsx.short_description = "Exportar selecionados para XLSX"

# Action para exportar saídas selecionadas para XLSX
def exportar_saidas_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Descrição', 'Valor', 'Data', 'Categoria', 'Pago por'])
    for obj in queryset:
        ws.append([
            obj.descricao,
            obj.valor,
            obj.data.strftime('%d/%m/%Y'),
            obj.categoria,
            obj.pago_por
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=saidas.xlsx'
    wb.save(response)
    return response

exportar_saidas_xlsx.short_description = "Exportar selecionados para XLSX"

# Action para exportar contas a receber selecionadas para XLSX
def exportar_contasreceber_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Descrição', 'Documento', 'Valor', 'Data Vencimento', 'Categoria', 'Recebido'])
    for obj in queryset:
        ws.append([
            obj.descricao,
            obj.documento,
            obj.valor,
            obj.data_vencimento.strftime('%d/%m/%Y'),
            obj.categoria,
            'Sim' if obj.recebido else 'Não'
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=contas_a_receber.xlsx'
    wb.save(response)
    return response

exportar_contasreceber_xlsx.short_description = "Exportar selecionados para XLSX"
import openpyxl
from django.http import HttpResponse
# Action para exportar contas a pagar selecionadas para XLSX
def exportar_contaspagar_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Descrição', 'Documento', 'Valor', 'Data Vencimento', 'Categoria', 'Pago'])
    for obj in queryset:
        ws.append([
            obj.descricao,
            obj.documento,
            obj.valor,
            obj.data_vencimento.strftime('%d/%m/%Y'),
            obj.categoria,
            'Sim' if obj.pago else 'Não'
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=contas_a_pagar.xlsx'
    wb.save(response)
    return response

exportar_contaspagar_xlsx.short_description = "Exportar selecionados para XLSX"
from django.contrib.admin import SimpleListFilter

# Filtros customizados por mês/ano
class MesAnoDataFilter(SimpleListFilter):
    title = 'Mês/Ano'
    parameter_name = 'mes_ano'

    def lookups(self, request, model_admin):
        datas = model_admin.model.objects.order_by('-data').values_list('data', flat=True)
        opcoes = set()
        for data in datas:
            opcoes.add((data.strftime('%Y-%m'), data.strftime('%m/%Y')))
        return sorted(opcoes, reverse=True)

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            ano, mes = value.split('-')
            return queryset.filter(data__year=ano, data__month=mes)
        return queryset

class MesAnoVencimentoPagarFilter(SimpleListFilter):
    title = 'Mês/Ano Vencimento'
    parameter_name = 'mes_ano_vencimento'

    def lookups(self, request, model_admin):
        datas = model_admin.model.objects.order_by('-data_vencimento').values_list('data_vencimento', flat=True)
        opcoes = set()
        for data in datas:
            opcoes.add((data.strftime('%Y-%m'), data.strftime('%m/%Y')))
        return sorted(opcoes, reverse=True)

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            ano, mes = value.split('-')
            return queryset.filter(data_vencimento__year=ano, data_vencimento__month=mes)
        return queryset

class MesAnoVencimentoFilter(SimpleListFilter):
    title = 'Mês/Ano Vencimento'
    parameter_name = 'mes_ano_vencimento'

    def lookups(self, request, model_admin):
        datas = model_admin.model.objects.order_by('-data_vencimento').values_list('data_vencimento', flat=True)
        opcoes = set()
        for data in datas:
            opcoes.add((data.strftime('%Y-%m'), data.strftime('%m/%Y')))
        return sorted(opcoes, reverse=True)

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            ano, mes = value.split('-')
            return queryset.filter(data_vencimento__year=ano, data_vencimento__month=mes)
        return queryset
from django.contrib import admin
from .models import Entrada, Saida, ContaPagar, ContaReceber
from .forms import ContaPagarForm, SaidaForm, EntradaForm, ContaReceberForm
from dateutil.relativedelta import relativedelta
from rangefilter.filters import DateRangeFilter


# Admins


@admin.register(Entrada)
class EntradaAdmin(admin.ModelAdmin):
    form = EntradaForm
    list_display = ('descricao', 'valor', 'data', 'categoria', 'atualizado_em')
    search_fields = ('descricao', 'categoria')
    list_filter = (MesAnoDataEntradaFilter, 'categoria')
    actions = [exportar_entradas_xlsx]

    def save_model(self, request, obj, form, change):
        parcelas = int(form.cleaned_data.get('parcelas', 1))
        if parcelas > 1:
            base_data = form.cleaned_data['data']
            for n in range(parcelas):
                novo = Entrada(
                    descricao=f"{obj.descricao} ({n+1}/{parcelas})",
                    valor=obj.valor,
                    data=base_data + relativedelta(months=n),
                    categoria=obj.categoria,
                    recebido_por=obj.recebido_por
                )
                novo.save()
        else:
            super().save_model(request, obj, form, change)



@admin.register(Saida)
class SaidaAdmin(admin.ModelAdmin):
    form = SaidaForm
    list_display = ('descricao', 'valor', 'data', 'categoria', 'atualizado_em')
    search_fields = ('descricao', 'categoria')
    list_filter = (MesAnoDataFilter, 'categoria')
    actions = [exportar_saidas_xlsx]

    def save_model(self, request, obj, form, change):
        parcelas = int(form.cleaned_data.get('parcelas', 1))
        if parcelas > 1:
            base_data = form.cleaned_data['data']
            for n in range(parcelas):
                novo = Saida(
                    descricao=f"{obj.descricao} ({n+1}/{parcelas})",
                    valor=obj.valor,
                    data=base_data + relativedelta(months=n),
                    categoria=obj.categoria,
                    pago_por=obj.pago_por
                )
                novo.save()
        else:
            super().save_model(request, obj, form, change)






@admin.register(ContaPagar)
class ContaPagarAdmin(admin.ModelAdmin):
    form = ContaPagarForm
    list_display = ('descricao','documento', 'valor', 'data_vencimento', 'categoria', 'pago')
    search_fields = ('descricao', 'categoria','documento')
    list_filter = (MesAnoVencimentoPagarFilter, 'categoria', 'pago')
    actions = [exportar_contaspagar_xlsx]

    def save_model(self, request, obj, form, change):
        parcelas = int(form.cleaned_data.get('parcelas', 1))
        if parcelas > 1:
            base_data = form.cleaned_data['data_vencimento']
            for n in range(parcelas):
                novo = ContaPagar(
                    descricao=f"{obj.descricao} ({n+1}/{parcelas})",
                    documento=obj.documento,
                    valor=obj.valor,
                    data_vencimento=base_data + relativedelta(months=n),
                    categoria=obj.categoria,
                    pago=obj.pago
                )
                novo.save()
        else:
            super().save_model(request, obj, form, change)



@admin.register(ContaReceber)
class ContaReceberAdmin(admin.ModelAdmin):
    form = ContaReceberForm
    list_display = ('descricao','documento', 'valor', 'data_vencimento', 'categoria', 'recebido')
    search_fields = ('descricao', 'categoria','documento')
    list_filter = (MesAnoVencimentoFilter, 'categoria', 'recebido')
    actions = [exportar_contasreceber_xlsx]

    def save_model(self, request, obj, form, change):
        parcelas = int(form.cleaned_data.get('parcelas', 1))
        if parcelas > 1:
            base_data = form.cleaned_data['data_vencimento']
            for n in range(parcelas):
                novo = ContaReceber(
                    descricao=f"{obj.descricao} ({n+1}/{parcelas})",
                    documento=obj.documento,
                    valor=obj.valor,
                    data_vencimento=base_data + relativedelta(months=n),
                    categoria=obj.categoria,
                    recebido=obj.recebido
                )
                novo.save()
        else:
            super().save_model(request, obj, form, change)

